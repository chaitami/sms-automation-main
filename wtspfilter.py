from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.common.by import By
from openpyxl import load_workbook, Workbook

# Initialize the webdriver for Firefox
driver = webdriver.Firefox()

wb = Workbook()
ws = wb.active

wberror = Workbook()
wserror = wberror.active
# Open the WhatsApp Web interface
driver.get("https://web.whatsapp.com/")

# Wait for the user to scan the QR code and log in
input("Press Enter after scanning the QR code and logging in…")
phone_numbers = []
phone_error = []
# Open the file containing the phone numbers

with open("phone.txt", "r") as f:
    for line in f:
        # Strip any whitespace from the line
        line = line.strip()
        # Navigate to the WhatsApp Web search page
        driver.get(f"https://web.whatsapp.com/send?phone={line}")
        #12
        time.sleep(12) # wait for the page to load
        # Check if the phone number is registered on WhatsApp
        
        
        if driver.find_elements("xpath","//div[contains(text(),'Phone number shared via url is invalid.')]"):
            print(f"{line} - Not registered on WhatsApp")
        elif driver.find_elements("xpath","//header/div[3]/div[1]/div[1]/div[1]/span[1]"):
            print(f"{line} - Registered on WhatsApp")
            phone_numbers.append(line)
        else:
            phone_error.append(line)
            print(f"{line} - ERROR")

for data in phone_numbers:
    ws.append([data])

for data1 in phone_error:
    wserror.append([data1])

wb.save(filename='checked.xlsx') 
wberror.save(filename='error.xlsx') 
# Close the webdriver
driver.quit()